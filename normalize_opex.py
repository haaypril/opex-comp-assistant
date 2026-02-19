#!/usr/bin/env python3
"""Build one Excel operating-expense comp workbook with monthly reconciliations."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


TARGET_CATEGORIES = [
    "payroll",
    "management fee",
    "general & admin & advertising",
    "repairs & maintenance",
    "insurance",
    "property taxes",
    "replacement reserves",
    "utilities",
]


@dataclass
class RuleMatch:
    category: str
    rule_name: str


def load_rules(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if {"ignore_patterns", "mapping_rules"} - set(payload.keys()):
        raise ValueError("Rules JSON must include ignore_patterns and mapping_rules")
    return payload


def load_overrides(path: Path | None) -> list[dict]:
    if not path or not path.exists():
        return []
    out: list[dict] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            out.append(
                {
                    "match_type": (row.get("match_type") or "").strip().lower(),
                    "pattern": (row.get("pattern") or "").strip(),
                    "category": (row.get("category") or "").strip().lower(),
                }
            )
    return out


def normalize_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", (value or "")).strip()


def parse_month_label(value) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    if not isinstance(value, str):
        return None
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(".", " ").replace("/", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text).strip()
    for fmt in ("%b %Y", "%B %Y", "%b %y", "%B %y", "%m %d %Y", "%m %d %y"):
        try:
            dt = datetime.strptime(text, fmt)
            year = dt.year + 2000 if dt.year < 100 else dt.year
            return f"{year:04d}-{dt.month:02d}"
        except ValueError:
            continue
    return None


def detect_month_columns(ws, max_cols: int) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=max_cols, values_only=True):
        for ci, v in enumerate(row, 1):
            month = parse_month_label(v)
            if month:
                mapping[ci] = month
    return mapping


def infer_property_name(ws, fallback: str) -> str:
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3, values_only=True):
        for cell in row:
            if isinstance(cell, str):
                t = normalize_text(cell)
                if t and re.search(r"[A-Za-z]", t):
                    return t
    return fallback


def derive_label_and_code(text_cells: list[tuple[int, str]]) -> tuple[str, str]:
    sorted_cells = sorted(text_cells, key=lambda x: x[0])
    first = sorted_cells[0][1]
    second = sorted_cells[1][1] if len(sorted_cells) > 1 else ""
    code = ""
    desc = first
    if re.fullmatch(r"[\d.]+", first):
        code = first
        desc = second or first
    elif " - " in first:
        a, b = first.split(" - ", 1)
        if re.fullmatch(r"[\d.]+", a.strip()):
            code = a.strip()
            desc = b.strip()
    return normalize_text(desc), normalize_text(code)


def should_ignore(description: str, ignore_patterns: list[str]) -> bool:
    d = description.lower()
    return any(re.search(p, d, re.IGNORECASE) for p in ignore_patterns)


def apply_overrides(description: str, account_code: str, overrides: list[dict]) -> RuleMatch | None:
    d = description.lower()
    code = (account_code or "").lower()
    for i, o in enumerate(overrides, 1):
        cat = o["category"]
        mt = o["match_type"]
        pat = o["pattern"].lower()
        if cat not in TARGET_CATEGORIES or not pat:
            continue
        if mt == "code" and code.startswith(pat):
            return RuleMatch(cat, f"override:{i}:code:{pat}")
        if mt == "contains" and pat in d:
            return RuleMatch(cat, f"override:{i}:contains:{pat}")
        if mt == "regex" and re.search(pat, d, re.IGNORECASE):
            return RuleMatch(cat, f"override:{i}:regex:{pat}")
    return None


def apply_rules(description: str, account_code: str, rules: list[dict]) -> RuleMatch | None:
    d = description.lower()
    code = (account_code or "").lower()
    for i, rule in enumerate(rules, 1):
        cat = (rule.get("category") or "").strip().lower()
        if cat not in TARGET_CATEGORIES:
            continue
        patterns = [str(x) for x in rule.get("patterns", [])]
        code_prefixes = [str(x).lower() for x in rule.get("code_prefixes", [])]
        excludes = [str(x) for x in rule.get("exclude_patterns", [])]
        if any(re.search(x, d, re.IGNORECASE) for x in excludes):
            continue
        if (code and any(code.startswith(p) for p in code_prefixes)) or any(
            re.search(p, d, re.IGNORECASE) for p in patterns
        ):
            return RuleMatch(cat, rule.get("name") or f"rule_{i}")
    return None


def detect_expense_window(ws, max_cols: int) -> tuple[int | None, int | None]:
    start_row = None
    end_row = None
    for r, row in enumerate(ws.iter_rows(min_row=1, max_col=max_cols, values_only=True), 1):
        text = " | ".join(normalize_text(v).lower() for v in row if isinstance(v, str) and normalize_text(v))
        if not text:
            continue
        if start_row is None and (
            (
                re.search(r"(^|\\| )expenses$", text) is not None
                and "revenue and expenses" not in text
            )
            or "operating expenses" in text
        ):
            start_row = r
            continue
        if start_row is not None and (
            "total operating expenses" in text
            or re.search(r"\btotal expenses\b", text)
        ):
            end_row = r
            break
    return start_row, end_row


def extract_reported_totals(ws, month_cols: dict[int, str], total_row: int | None) -> dict[str, float]:
    if total_row is None:
        return {}
    values = {}
    cells = next(
        ws.iter_rows(min_row=total_row, max_row=total_row, min_col=1, max_col=max(month_cols.keys(), default=1), values_only=True)
    )
    for ci, month in month_cols.items():
        if ci - 1 < len(cells):
            v = cells[ci - 1]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                values[month] = round(float(v), 2)
    return values


def extract_file_rows(path: Path, sheet_name: str | None) -> tuple[list[dict], dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    max_cols = min(max(ws.max_column or 30, 15), 40)
    month_cols = detect_month_columns(ws, max_cols)
    start_row, end_row = detect_expense_window(ws, max_cols)
    property_name = infer_property_name(ws, path.stem)
    reported_totals = extract_reported_totals(ws, month_cols, end_row)

    rows: list[dict] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_col=max_cols, values_only=True), 1):
        if start_row is not None and row_number <= start_row:
            continue
        if end_row is not None and row_number >= end_row:
            continue

        text_cells = []
        for ci, v in enumerate(row, 1):
            if isinstance(v, str):
                t = normalize_text(v)
                if t:
                    text_cells.append((ci, t))
        if not text_cells:
            continue
        description, code = derive_label_and_code(text_cells)
        if not description or description.lower().startswith("total "):
            continue

        for ci, month in month_cols.items():
            if ci - 1 >= len(row):
                continue
            value = row[ci - 1]
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                rows.append(
                    {
                        "property_name": property_name,
                        "file_name": path.name,
                        "sheet_name": ws.title,
                        "row_number": row_number,
                        "account_code": code,
                        "description": description,
                        "month": month,
                        "amount": float(value),
                    }
                )
    meta = {"property_name": property_name, "file_name": path.name, "reported_monthly_totals": reported_totals}
    return rows, meta


def classify_rows(rows: list[dict], rules_payload: dict, overrides: list[dict]) -> tuple[list[dict], list[dict]]:
    normalized_rows = []
    unmapped_rows = []
    ignores = rules_payload["ignore_patterns"]
    rules = rules_payload["mapping_rules"]

    for row in rows:
        desc = row["description"]
        code = row["account_code"]
        if should_ignore(desc, ignores):
            continue
        match = apply_overrides(desc, code, overrides) or apply_rules(desc, code, rules)
        base = {
            "property_name": row["property_name"],
            "file_name": row["file_name"],
            "sheet_name": row["sheet_name"],
            "row_number": row["row_number"],
            "account_code": code,
            "description": desc,
            "month": row["month"],
            "amount": round(row["amount"], 2),
        }
        if match:
            normalized_rows.append({**base, "normalized_category": match.category, "matched_rule": match.rule_name})
        else:
            unmapped_rows.append(base)
    return normalized_rows, unmapped_rows


def build_monthly_comp(normalized_rows: list[dict]) -> list[dict]:
    agg = defaultdict(float)
    for r in normalized_rows:
        key = (r["property_name"], r["file_name"], r["month"], r["normalized_category"])
        agg[key] += float(r["amount"])
    row_keys = sorted({(k[0], k[1], k[2]) for k in agg.keys()}, key=lambda x: (x[0].lower(), x[2]))
    out = []
    for property_name, file_name, month in row_keys:
        row = {"property_name": property_name, "file_name": file_name, "month": month}
        for cat in TARGET_CATEGORIES:
            row[cat] = round(agg.get((property_name, file_name, month, cat), 0.0), 2)
        out.append(row)
    return out


def build_reconciliation(monthly_rows: list[dict], source_meta: list[dict]) -> list[dict]:
    norm = {}
    for r in monthly_rows:
        total = sum(float(r[c]) for c in TARGET_CATEGORIES)
        norm[(r["file_name"], r["month"])] = round(total, 2)

    out = []
    for meta in source_meta:
        file_name = meta["file_name"]
        property_name = meta["property_name"]
        for month in sorted(meta["reported_monthly_totals"].keys()):
            source_total = round(float(meta["reported_monthly_totals"][month]), 2)
            normalized_total = norm.get((file_name, month), 0.0)
            out.append(
                {
                    "property_name": property_name,
                    "file_name": file_name,
                    "month": month,
                    "source_total_expenses": source_total,
                    "normalized_total_expenses": normalized_total,
                    "variance": round(normalized_total - source_total, 2),
                }
            )
    return out


def format_month_label(month: str) -> str:
    year, mon = month.split("-")
    return f"{int(mon)}/1/{str(year)[2:]}"


def month_to_date(month: str) -> date:
    year, mon = month.split("-")
    return date(int(year), int(mon), 1)


def write_excel(path: Path, monthly_rows: list[dict], recon_rows: list[dict], normalized_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comp"

    property_names = sorted({r["property_name"] for r in monthly_rows})
    months = sorted({r["month"] for r in monthly_rows})
    value_lookup = {}
    for r in monthly_rows:
        for cat in TARGET_CATEGORIES:
            value_lookup[(r["month"], r["property_name"], cat)] = float(r[cat])

    row_ptr = 1
    ws.cell(row=row_ptr, column=1, value="Monthly Comp (Categories Down, Deals Across)")
    row_ptr += 2

    for month in months:
        month_cell = ws.cell(row=row_ptr, column=1, value=month_to_date(month))
        month_cell.number_format = "m/d/yy"
        row_ptr += 1
        ws.cell(row=row_ptr, column=1, value="expense_category")
        for col_idx, property_name in enumerate(property_names, start=2):
            ws.cell(row=row_ptr, column=col_idx, value=property_name)
        row_ptr += 1

        first_data_row = row_ptr
        for cat in TARGET_CATEGORIES:
            ws.cell(row=row_ptr, column=1, value=cat)
            for col_idx, property_name in enumerate(property_names, start=2):
                c = ws.cell(
                    row=row_ptr,
                    column=col_idx,
                    value=round(value_lookup.get((month, property_name, cat), 0.0), 2),
                )
                c.number_format = ACCOUNTING_FORMAT
            start_col = openpyxl.utils.get_column_letter(2)
            end_col = openpyxl.utils.get_column_letter(len(property_names) + 1)
            row_ptr += 1

        total_row = row_ptr
        ws.cell(row=total_row, column=1, value="total_opex")
        for col_idx in range(2, len(property_names) + 2):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            c = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}{first_data_row}:{col_letter}{total_row-1})",
            )
            c.number_format = ACCOUNTING_FORMAT
        row_ptr += 2

    ws.cell(row=row_ptr, column=1, value="Reconciliation Check")
    row_ptr += 1
    recon_header = [
        "property_name",
        "file_name",
        "month",
        "source_total_expenses",
        "normalized_total_expenses",
        "variance",
    ]
    for c, h in enumerate(recon_header, start=1):
        ws.cell(row=row_ptr, column=c, value=h)
    row_ptr += 1
    for r in recon_rows:
        for c, h in enumerate(recon_header, start=1):
            cell = ws.cell(row=row_ptr, column=c, value=r[h])
            if h == "month" and isinstance(r[h], str) and re.match(r"^\d{4}-\d{2}$", r[h]):
                cell.value = month_to_date(r[h])
                cell.number_format = "m/d/yy"
            if h in {"source_total_expenses", "normalized_total_expenses", "variance"}:
                cell.number_format = ACCOUNTING_FORMAT
        row_ptr += 1

    row_ptr += 2
    ws.cell(row=row_ptr, column=1, value="Individual Normalized Line Items")
    row_ptr += 1
    line_header = [
        "property_name",
        "file_name",
        "sheet_name",
        "row_number",
        "account_code",
        "description",
        "month",
        "amount",
        "normalized_category",
        "matched_rule",
    ]
    for c, h in enumerate(line_header, start=1):
        ws.cell(row=row_ptr, column=c, value=h)
    row_ptr += 1
    for r in normalized_rows:
        for c, h in enumerate(line_header, start=1):
            cell = ws.cell(row=row_ptr, column=c, value=r[h])
            if h == "month" and isinstance(r[h], str) and re.match(r"^\d{4}-\d{2}$", r[h]):
                cell.value = month_to_date(r[h])
                cell.number_format = "m/d/yy"
            if h == "amount":
                cell.number_format = ACCOUNTING_FORMAT
        row_ptr += 1
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Normalize op-ex into one Excel workbook.")
    p.add_argument("inputs", nargs="+", help="Input workbook files (.xlsx/.xlsm).")
    p.add_argument("--rules", default="config/expense_mapping_rules.json")
    p.add_argument("--overrides", default="config/expense_overrides.csv")
    p.add_argument("--output", default="output/opex_normalized/opex_portfolio_comp.xlsx")
    p.add_argument("--sheet", default=None)
    return p.parse_args()


def main() -> None:
    args = parse_args()
    inputs = [Path(x).expanduser() for x in args.inputs]
    for p in inputs:
        if not p.exists():
            raise FileNotFoundError(f"Input file not found: {p}")

    rules = load_rules(Path(args.rules))
    overrides = load_overrides(Path(args.overrides))

    extracted = []
    source_meta = []
    for p in inputs:
        rows, meta = extract_file_rows(p, args.sheet)
        extracted.extend(rows)
        source_meta.append(meta)

    normalized, unmapped = classify_rows(extracted, rules, overrides)
    monthly_rows = build_monthly_comp(normalized)
    recon_rows = build_reconciliation(monthly_rows, source_meta)
    write_excel(Path(args.output), monthly_rows, recon_rows, normalized)

    bad = [r for r in recon_rows if abs(float(r["variance"])) > 0.01]
    print(f"Processed monthly line items: {len(extracted)}")
    print(f"Mapped monthly line items:    {len(normalized)}")
    print(f"Unmapped monthly line items:  {len(unmapped)}")
    print(f"Property-month rows:          {len(monthly_rows)}")
    print(f"Reconciliation rows:          {len(recon_rows)}")
    print(f"Rows with variance != 0:      {len(bad)}")
    print(f"Workbook:                     {Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
